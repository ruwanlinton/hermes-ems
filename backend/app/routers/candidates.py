import csv
import io
from datetime import date
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import get_db
from sqlalchemy.orm import selectinload

from app.db.models import Candidate, BatchMembership, Exam, Result, Subject, Examination, User
from app.auth.jwt import require_roles
from app.schemas.candidate import CandidateCreate, CandidateUpdate, CandidateOut, ImportResult, ImportError
from app.schemas.stats import CandidatePerformance, CandidateInfo, CandidatePaperResult, CandidateExaminationResult

router = APIRouter()

_any_role = require_roles("admin", "creator", "marker", "viewer")
_creator_plus = require_roles("admin", "creator")


# ---------------------------------------------------------------------------
# List / search (paginated)
# ---------------------------------------------------------------------------

@router.get("/candidates", response_model=dict)
async def list_candidates(
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_any_role),
):
    q = select(Candidate)
    if search:
        pattern = f"%{search}%"
        q = q.where(
            or_(
                Candidate.name.ilike(pattern),
                Candidate.registration_number.ilike(pattern),
            )
        )
    total_result = await db.execute(select(func.count()).select_from(q.subquery()))
    total = total_result.scalar_one()

    q = q.order_by(Candidate.name).offset((page - 1) * page_size).limit(page_size)
    rows = (await db.execute(q)).scalars().all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [CandidateOut.model_validate(r) for r in rows],
    }


# ---------------------------------------------------------------------------
# Create
# ---------------------------------------------------------------------------

@router.post("/candidates", response_model=CandidateOut, status_code=status.HTTP_201_CREATED)
async def create_candidate(
    payload: CandidateCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    existing = (await db.execute(
        select(Candidate).where(Candidate.registration_number == payload.registration_number)
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="Registration number already exists")

    candidate = Candidate(**payload.model_dump())
    db.add(candidate)
    await db.commit()
    await db.refresh(candidate)
    return candidate


# ---------------------------------------------------------------------------
# Get single
# ---------------------------------------------------------------------------

@router.get("/candidates/{candidate_id}", response_model=CandidateOut)
async def get_candidate(
    candidate_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_any_role),
):
    candidate = (await db.execute(
        select(Candidate).where(Candidate.id == candidate_id)
    )).scalar_one_or_none()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    return candidate


# ---------------------------------------------------------------------------
# Update
# ---------------------------------------------------------------------------

@router.patch("/candidates/{candidate_id}", response_model=CandidateOut)
async def update_candidate(
    candidate_id: str,
    payload: CandidateUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    candidate = (await db.execute(
        select(Candidate).where(Candidate.id == candidate_id)
    )).scalar_one_or_none()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    updates = payload.model_dump(exclude_unset=True)
    if "registration_number" in updates and updates["registration_number"] != candidate.registration_number:
        clash = (await db.execute(
            select(Candidate).where(Candidate.registration_number == updates["registration_number"])
        )).scalar_one_or_none()
        if clash:
            raise HTTPException(status_code=409, detail="Registration number already exists")

    for field, value in updates.items():
        setattr(candidate, field, value)
    await db.commit()
    await db.refresh(candidate)
    return candidate


# ---------------------------------------------------------------------------
# Delete
# ---------------------------------------------------------------------------

@router.delete("/candidates/{candidate_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_candidate(
    candidate_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    candidate = (await db.execute(
        select(Candidate).where(Candidate.id == candidate_id)
    )).scalar_one_or_none()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    enrollment_count = (await db.execute(
        select(func.count()).where(BatchMembership.candidate_id == candidate_id)
    )).scalar_one()
    if enrollment_count > 0:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot delete: candidate is enrolled in {enrollment_count} batch(es). Unenroll first.",
        )

    await db.delete(candidate)
    await db.commit()


# ---------------------------------------------------------------------------
# Import (CSV or XLSX)
# ---------------------------------------------------------------------------

def _parse_date(value: str) -> Optional[date]:
    if not value or not value.strip():
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return date.fromisoformat(value.strip()) if fmt == "%Y-%m-%d" else date(*[int(x) for x in value.strip().split(fmt[2])])
        except Exception:
            pass
    try:
        from datetime import datetime as dt
        return dt.strptime(value.strip(), "%Y-%m-%d").date()
    except Exception:
        return None


def _parse_date_robust(value: str) -> Optional[date]:
    """Try several common date formats and return a date or None."""
    if not value or not str(value).strip():
        return None
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            from datetime import datetime as dt
            return dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


@router.post("/candidates/import", response_model=ImportResult)
async def import_candidates(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []

    if filename.endswith(".xlsx") or file.content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        wb.close()
    else:
        # Treat as CSV
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower(): v.strip() for k, v in row.items()})

    imported = 0
    updated = 0
    errors: list[ImportError] = []

    for i, row in enumerate(rows, start=2):  # row 1 = header
        reg = row.get("registration_number", "").strip()
        name = row.get("name", "").strip()

        if not reg:
            errors.append(ImportError(row=i, message="registration_number is required"))
            continue
        if not name:
            errors.append(ImportError(row=i, message="name is required"))
            continue

        dob = _parse_date_robust(row.get("date_of_birth", ""))
        address = row.get("address", "") or None
        mobile = row.get("mobile", "") or None

        existing = (await db.execute(
            select(Candidate).where(Candidate.registration_number == reg)
        )).scalar_one_or_none()

        if existing:
            existing.name = name
            existing.date_of_birth = dob
            existing.address = address
            existing.mobile = mobile
            updated += 1
        else:
            db.add(Candidate(
                registration_number=reg,
                name=name,
                date_of_birth=dob,
                address=address,
                mobile=mobile,
            ))
            imported += 1

    await db.commit()
    return ImportResult(imported=imported, updated=updated, errors=errors)


# ---------------------------------------------------------------------------
# Export (CSV or XLSX)
# ---------------------------------------------------------------------------

@router.get("/candidates/export")
async def export_candidates(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_any_role),
):
    q = select(Candidate)
    if search:
        pattern = f"%{search}%"
        q = q.where(
            or_(
                Candidate.name.ilike(pattern),
                Candidate.registration_number.ilike(pattern),
            )
        )
    candidates = (await db.execute(q.order_by(Candidate.name))).scalars().all()

    headers = ["registration_number", "name", "date_of_birth", "address", "mobile"]

    def row(c: Candidate):
        return [
            c.registration_number,
            c.name,
            c.date_of_birth.isoformat() if c.date_of_birth else "",
            c.address or "",
            c.mobile or "",
        ]

    if format == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidates"
        ws.append(headers)
        for c in candidates:
            ws.append(row(c))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=candidates.xlsx"},
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for c in candidates:
            writer.writerow(row(c))
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=candidates.csv"},
        )


# ---------------------------------------------------------------------------
# Candidate performance (cross-examination)
# ---------------------------------------------------------------------------

@router.get("/candidates/{candidate_id}/performance", response_model=CandidatePerformance)
async def get_candidate_performance(
    candidate_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_any_role),
):
    candidate = (await db.execute(
        select(Candidate).where(Candidate.id == candidate_id)
    )).scalar_one_or_none()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    # Load all results for this candidate, with full exam → subject → examination chain
    results = (await db.execute(
        select(Result)
        .options(
            selectinload(Result.exam)
            .selectinload(Exam.subject)
            .selectinload(Subject.examination)
        )
        .where(Result.candidate_id == candidate_id)
        .order_by(Result.percentage.desc())
    )).scalars().all()

    # Group results by examination
    by_examination: dict[str, dict] = {}
    for r in results:
        exam = r.exam
        if not exam or not exam.subject or not exam.subject.examination:
            continue
        examination = exam.subject.examination
        eid = examination.id
        if eid not in by_examination:
            by_examination[eid] = {"examination": examination, "papers": []}
        by_examination[eid]["papers"].append(CandidatePaperResult(
            paper_id=exam.id,
            title=exam.title,
            score=r.score,
            percentage=r.percentage,
            passed=r.percentage >= exam.pass_mark,
            pass_mark=exam.pass_mark,
        ))

    examinations = []
    for eid, data in by_examination.items():
        papers = data["papers"]
        overall = round(sum(p.percentage for p in papers) / len(papers), 2) if papers else 0.0
        examinations.append(CandidateExaminationResult(
            examination_id=eid,
            title=data["examination"].title,
            papers=papers,
            overall_percentage=overall,
        ))

    return CandidatePerformance(
        candidate=CandidateInfo(
            id=candidate.id,
            registration_number=candidate.registration_number,
            name=candidate.name,
        ),
        examinations=examinations,
    )
