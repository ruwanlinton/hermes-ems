import csv
import io
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db.session import get_db
from app.db.models import Batch, BatchMembership, Candidate, Examination, User
from app.auth.jwt import require_roles
from app.schemas.batch import (
    BatchCreate,
    BatchUpdate,
    BatchOut,
    MembershipCreate,
    MembershipOut,
    MembershipImportResult,
    MembershipImportError,
)

router = APIRouter()

_any_role = require_roles("admin", "creator", "marker", "viewer")
_creator_plus = require_roles("admin", "creator")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _get_examination_or_404(eid: str, db: AsyncSession) -> Examination:
    exam = (await db.execute(
        select(Examination).where(Examination.id == eid)
    )).scalar_one_or_none()
    if not exam:
        raise HTTPException(status_code=404, detail="Examination not found")
    return exam


async def _get_batch_or_404(eid: str, bid: str, db: AsyncSession) -> Batch:
    batch = (await db.execute(
        select(Batch).where(Batch.id == bid, Batch.examination_id == eid)
    )).scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    return batch


def _assert_not_closed(examination: Examination, action: str):
    if examination.status == "closed":
        raise HTTPException(
            status_code=409,
            detail=f"Examination is closed: {action} is not permitted.",
        )


async def _batch_out(batch: Batch, db: AsyncSession) -> BatchOut:
    count = (await db.execute(
        select(func.count()).where(BatchMembership.batch_id == batch.id)
    )).scalar_one()
    return BatchOut(
        id=batch.id,
        examination_id=batch.examination_id,
        name=batch.name,
        member_count=count,
        created_at=batch.created_at,
    )


def _membership_out(m: BatchMembership) -> MembershipOut:
    return MembershipOut(
        id=m.id,
        batch_id=m.batch_id,
        candidate_id=m.candidate_id,
        index_number=m.index_number,
        candidate_name=m.candidate.name,
        candidate_registration_number=m.candidate.registration_number,
    )


# ---------------------------------------------------------------------------
# Batches CRUD
# ---------------------------------------------------------------------------

@router.get("/examinations/{eid}/batches", response_model=list[BatchOut])
async def list_batches(
    eid: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_any_role),
):
    await _get_examination_or_404(eid, db)
    batches = (await db.execute(
        select(Batch).where(Batch.examination_id == eid).order_by(Batch.name)
    )).scalars().all()
    return [await _batch_out(b, db) for b in batches]


@router.post("/examinations/{eid}/batches", response_model=BatchOut, status_code=status.HTTP_201_CREATED)
async def create_batch(
    eid: str,
    payload: BatchCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    examination = await _get_examination_or_404(eid, db)
    _assert_not_closed(examination, "creating batches")

    clash = (await db.execute(
        select(Batch).where(Batch.examination_id == eid, Batch.name == payload.name)
    )).scalar_one_or_none()
    if clash:
        raise HTTPException(status_code=409, detail="A batch with this name already exists.")

    batch = Batch(examination_id=eid, name=payload.name)
    db.add(batch)
    await db.commit()
    await db.refresh(batch)
    return await _batch_out(batch, db)


@router.get("/examinations/{eid}/batches/{bid}", response_model=BatchOut)
async def get_batch(
    eid: str,
    bid: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_any_role),
):
    batch = await _get_batch_or_404(eid, bid, db)
    return await _batch_out(batch, db)


@router.patch("/examinations/{eid}/batches/{bid}", response_model=BatchOut)
async def update_batch(
    eid: str,
    bid: str,
    payload: BatchUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    examination = await _get_examination_or_404(eid, db)
    _assert_not_closed(examination, "renaming batches")
    batch = await _get_batch_or_404(eid, bid, db)

    if payload.name != batch.name:
        clash = (await db.execute(
            select(Batch).where(Batch.examination_id == eid, Batch.name == payload.name)
        )).scalar_one_or_none()
        if clash:
            raise HTTPException(status_code=409, detail="A batch with this name already exists.")

    batch.name = payload.name
    await db.commit()
    await db.refresh(batch)
    return await _batch_out(batch, db)


@router.delete("/examinations/{eid}/batches/{bid}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_batch(
    eid: str,
    bid: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    examination = await _get_examination_or_404(eid, db)
    _assert_not_closed(examination, "deleting batches")
    batch = await _get_batch_or_404(eid, bid, db)

    member_count = (await db.execute(
        select(func.count()).where(BatchMembership.batch_id == bid)
    )).scalar_one()
    if member_count > 0:
        raise HTTPException(
            status_code=409,
            detail="Cannot delete: batch has members. Unenroll all members first.",
        )

    await db.delete(batch)
    await db.commit()


# ---------------------------------------------------------------------------
# Members
# ---------------------------------------------------------------------------

@router.get("/examinations/{eid}/batches/{bid}/members", response_model=dict)
async def list_members(
    eid: str,
    bid: str,
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_any_role),
):
    await _get_batch_or_404(eid, bid, db)

    q = (
        select(BatchMembership)
        .options(selectinload(BatchMembership.candidate))
        .where(BatchMembership.batch_id == bid)
    )
    if search:
        pattern = f"%{search}%"
        q = q.join(Candidate, Candidate.id == BatchMembership.candidate_id).where(
            (Candidate.name.ilike(pattern))
            | (Candidate.registration_number.ilike(pattern))
            | (BatchMembership.index_number.ilike(pattern))
        )

    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar_one()
    rows = (await db.execute(
        q.order_by(BatchMembership.index_number).offset((page - 1) * page_size).limit(page_size)
    )).scalars().all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [_membership_out(m) for m in rows],
    }


@router.post("/examinations/{eid}/batches/{bid}/members", response_model=MembershipOut, status_code=status.HTTP_201_CREATED)
async def enroll_member(
    eid: str,
    bid: str,
    payload: MembershipCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    examination = await _get_examination_or_404(eid, db)
    _assert_not_closed(examination, "enrolling candidates")
    await _get_batch_or_404(eid, bid, db)

    # Verify candidate exists
    candidate = (await db.execute(
        select(Candidate).where(Candidate.id == payload.candidate_id)
    )).scalar_one_or_none()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    # Check duplicate candidate in batch
    dup_candidate = (await db.execute(
        select(BatchMembership).where(
            BatchMembership.batch_id == bid,
            BatchMembership.candidate_id == payload.candidate_id,
        )
    )).scalar_one_or_none()
    if dup_candidate:
        raise HTTPException(status_code=409, detail="Candidate is already enrolled in this batch.")

    # Check duplicate index number in batch
    dup_index = (await db.execute(
        select(BatchMembership).where(
            BatchMembership.batch_id == bid,
            BatchMembership.index_number == payload.index_number,
        )
    )).scalar_one_or_none()
    if dup_index:
        raise HTTPException(status_code=409, detail="Index number is already assigned in this batch.")

    membership = BatchMembership(
        batch_id=bid,
        candidate_id=payload.candidate_id,
        index_number=payload.index_number,
    )
    db.add(membership)
    await db.commit()
    await db.refresh(membership)

    # Re-fetch with candidate loaded
    m = (await db.execute(
        select(BatchMembership)
        .options(selectinload(BatchMembership.candidate))
        .where(BatchMembership.id == membership.id)
    )).scalar_one()
    return _membership_out(m)


@router.post("/examinations/{eid}/batches/{bid}/members/import", response_model=MembershipImportResult)
async def import_members(
    eid: str,
    bid: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    examination = await _get_examination_or_404(eid, db)
    _assert_not_closed(examination, "enrolling candidates")
    await _get_batch_or_404(eid, bid, db)

    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []
    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        wb.close()
    else:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower(): v.strip() for k, v in row.items()})

    enrolled = 0
    errors: list[MembershipImportError] = []

    for i, row in enumerate(rows, start=2):
        reg = row.get("registration_number", "").strip()
        index_number = row.get("index_number", "").strip()

        if not reg:
            errors.append(MembershipImportError(row=i, message="registration_number is required"))
            continue
        if not index_number:
            errors.append(MembershipImportError(row=i, message="index_number is required"))
            continue

        candidate = (await db.execute(
            select(Candidate).where(Candidate.registration_number == reg)
        )).scalar_one_or_none()
        if not candidate:
            errors.append(MembershipImportError(row=i, message=f"No candidate with registration_number '{reg}'"))
            continue

        dup_candidate = (await db.execute(
            select(BatchMembership).where(
                BatchMembership.batch_id == bid,
                BatchMembership.candidate_id == candidate.id,
            )
        )).scalar_one_or_none()
        if dup_candidate:
            errors.append(MembershipImportError(row=i, message=f"'{reg}' is already enrolled in this batch"))
            continue

        dup_index = (await db.execute(
            select(BatchMembership).where(
                BatchMembership.batch_id == bid,
                BatchMembership.index_number == index_number,
            )
        )).scalar_one_or_none()
        if dup_index:
            errors.append(MembershipImportError(row=i, message=f"Index number '{index_number}' already assigned"))
            continue

        db.add(BatchMembership(batch_id=bid, candidate_id=candidate.id, index_number=index_number))
        enrolled += 1

    await db.commit()
    return MembershipImportResult(enrolled=enrolled, errors=errors)


@router.delete("/examinations/{eid}/batches/{bid}/members/{mid}", status_code=status.HTTP_204_NO_CONTENT)
async def unenroll_member(
    eid: str,
    bid: str,
    mid: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_creator_plus),
):
    examination = await _get_examination_or_404(eid, db)
    _assert_not_closed(examination, "unenrolling candidates")

    membership = (await db.execute(
        select(BatchMembership).where(BatchMembership.id == mid, BatchMembership.batch_id == bid)
    )).scalar_one_or_none()
    if not membership:
        raise HTTPException(status_code=404, detail="Membership not found")

    await db.delete(membership)
    await db.commit()


@router.get("/examinations/{eid}/batches/{bid}/members/export")
async def export_members(
    eid: str,
    bid: str,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(_any_role),
):
    await _get_batch_or_404(eid, bid, db)
    members = (await db.execute(
        select(BatchMembership)
        .options(selectinload(BatchMembership.candidate))
        .where(BatchMembership.batch_id == bid)
        .order_by(BatchMembership.index_number)
    )).scalars().all()

    headers = ["index_number", "registration_number", "name"]

    def row(m: BatchMembership):
        return [m.index_number, m.candidate.registration_number, m.candidate.name]

    if format == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        ws.append(headers)
        for m in members:
            ws.append(row(m))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=members.xlsx"},
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for m in members:
            writer.writerow(row(m))
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=members.csv"},
        )
